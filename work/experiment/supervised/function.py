import os
#os.environ["OMP_NUM_THREADS"] = "1"
#os.environ["MKL_NUM_THREADS"] = "1"
#os.environ["OPENBLAS_NUM_THREADS"] = "1"
#os.environ["NUMEXPR_NUM_THREADS"] = "1"
#os.environ['PYTHONHASHSEED'] = '42'

import glob
import json
import time
from abc import ABC, abstractmethod

import numpy as np
import pandas as pd
import torch
import torch.nn.functional as F
import plotly.graph_objects as go
from pathlib import Path
import optuna

from sklearn.metrics import (
    average_precision_score,
    roc_curve,
    f1_score,
    precision_recall_curve,
    recall_score,
    precision_score,
    confusion_matrix,
)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from plotly.subplots import make_subplots
import random
from sklearn.preprocessing import RobustScaler
from sktime.classification.deep_learning.mlp import MLPClassifier

#def set_seed(seed=42):
    #os.environ['PYTHONHASHSEED'] = str(seed)
    #os.environ["OMP_NUM_THREADS"] = "1"
    #os.environ["MKL_NUM_THREADS"] = "1"
    #os.environ["OPENBLAS_NUM_THREADS"] = "1"
    #os.environ["NUMEXPR_NUM_THREADS"] = "1"
    #random.seed(seed)
    #np.random.seed(seed)
    #torch.manual_seed(seed)
    #if torch.cuda.is_available():
    #    torch.cuda.manual_seed(seed)
    #    torch.cuda.manual_seed_all(seed)
    #torch.use_deterministic_algorithms(True)
    #torch.backends.cudnn.deterministic = True
    #torch.backends.cudnn.benchmark = False


class MetricsBase:
    @staticmethod
    def event_lengths(events):
        durations = []
        for start, end in events:
            start = pd.to_datetime(start)
            end = pd.to_datetime(end)
            durations.append((end - start).total_seconds() / 60)

        if not durations:
            return {"min": 0.0, "median": 0.0}

        return {
            "min": int(np.min(durations)),
            "median": int(np.median(durations)),
        }
    
    @staticmethod
    def get_anomaly_lengths(data, default_value=0):
        target = data["target"].values

        lengths = []
        current_len = 0

        for val in target:
            if val == 1:
                current_len += 1
            else:
                if current_len > 0:
                    lengths.append(current_len)
                    current_len = 0

        if current_len > 0:
            lengths.append(current_len)

        if not lengths:
            return {"min": 0.0, "median": 0.0}

        return {
            "min": int(np.min(lengths)),
            "median": int(np.median(lengths)),
        }

    @staticmethod
    def get_anomaly_lengths_time(data):
        if isinstance(data, pd.Series):
            target = data.copy()
        else:
            target = data['target'].copy()

        target.index = pd.to_datetime(target.index)

        lengths = []
        current_start = None

        for t, val in target.items():
            if val == 1:
                if current_start is None:
                    current_start = t
            else:
                if current_start is not None:
                    duration = (t - current_start).total_seconds() / 60
                    lengths.append(duration)
                    current_start = None

        # если аномалия до конца
        if current_start is not None:
            duration = (target.index[-1] - current_start).total_seconds() / 60
            lengths.append(duration)

        if not lengths:
            return {"min": 0.0, "median": 0.0}

        return {
            "min": int(np.min(lengths)),
            "median": int(np.median(lengths)),
        }

    @staticmethod
    def event_params(min_event_len, median_event_len):
        def clip(value, vmin, vmax):
            return max(vmin, min(value, vmax))

        gap_fill = clip(min_event_len, 3, 15)
        min_event = clip(min_event_len * 0.5, 3, 120)
        event_interval = clip(median_event_len * 4 / 5, 20, 300)
        window_size = clip(median_event_len * 0.4, 5, 120)

        return int(gap_fill), int(min_event), int(event_interval), int(window_size)

    @staticmethod
    def pr_auc(y_true, scores):
        return average_precision_score(np.asarray(y_true), np.asarray(scores))

    @staticmethod
    def binary_to_events(y_bin: pd.Series, gap_minutes=10, min_len_minutes=1):
        y_bin = y_bin.sort_index().astype(int)
        ones = y_bin[y_bin == 1].dropna()
        if ones.empty:
            return []

        gap = pd.Timedelta(minutes=gap_minutes)
        min_len = pd.Timedelta(minutes=min_len_minutes)

        times = ones.index
        events = []
        start = prev = times[0]

        for t in times[1:]:
            if (t - prev) <= gap:
                prev = t
            else:
                end = prev
                if (end - start) >= min_len:
                    events.append((start, end))
                start = prev = t

        end = prev
        if (end - start) >= min_len:
            events.append((start, end))

        return events

    @staticmethod
    def f1_at_fpr(y_true, scores, target_fpr=0.01):
        y_true = np.asarray(y_true)
        scores = np.asarray(scores)

        fpr, _, thresholds = roc_curve(y_true, scores)
        valid_idx = np.where(fpr <= target_fpr)[0]
        if len(valid_idx) == 0:
            return 0.0

        threshold = thresholds[valid_idx[-1]]
        y_pred = (scores >= threshold).astype(int)
        return f1_score(y_true, y_pred, zero_division=0)

    @staticmethod
    def precision_at_k(y_true, scores, k):
        y_true = np.asarray(y_true)
        scores = np.asarray(scores)

        if k <= 0:
            raise ValueError("k must be > 0")
        if len(y_true) == 0:
            return 0.0

        k = min(k, len(y_true))
        top_k_idx = np.argsort(scores)[::-1][:k]
        return float(y_true[top_k_idx].sum() / k)

    @staticmethod
    def recall_at_k(y_true, scores, k):
        y_true = np.asarray(y_true)
        scores = np.asarray(scores)

        if k <= 0:
            raise ValueError("k must be > 0")

        total_anomalies = y_true.sum()
        if total_anomalies == 0:
            return 0.0

        k = min(k, len(y_true))
        top_k_idx = np.argsort(scores)[::-1][:k]
        found_anomalies = y_true[top_k_idx].sum()

        return float(found_anomalies / total_anomalies)

    @staticmethod
    def _inside_true_event(pred_start, pred_end, true_start, true_end):
        return true_start <= pred_start and pred_end <= true_end

    @staticmethod
    def _is_timely_detection(pred_start, true_start, max_min=180):
        window = pd.Timedelta(minutes=max_min)
        return (true_start - window) <= pred_start <= (true_start + window)

    @classmethod
    def event_recall(cls, true_events, pred_events, max_min=180):
        if len(true_events) == 0:
            return 0.0

        detected = 0
        for t_start, _ in true_events:
            found = False
            for p_start, _ in pred_events:
                if cls._is_timely_detection(p_start, t_start, max_min=max_min):
                    found = True
                    break
            if found:
                detected += 1

        return detected / len(true_events)

    @classmethod
    def event_precision(cls, true_events, pred_events, max_min=180):
        if len(pred_events) == 0:
            return 0.0

        correct_pred_idx = set()

        for t_start, t_end in true_events:
            for i, (p_start, p_end) in enumerate(pred_events):
                if cls._is_timely_detection(p_start, t_start, max_min=max_min):
                    correct_pred_idx.add(i)
                elif cls._inside_true_event(p_start, p_end, t_start, t_end):
                    correct_pred_idx.add(i)

        return len(correct_pred_idx) / len(pred_events)

    @classmethod
    def time_to_detect(cls, true_events, pred_events, reduction="mean", unit="minutes", max_min=180, absolute=False):
        window = pd.Timedelta(minutes=max_min)
        delays = []

        for t_start, _ in true_events:
            valid_starts = []
            for p_start, _ in pred_events:
                if (t_start - window) <= p_start <= (t_start + window):
                    valid_starts.append(p_start)

            if valid_starts:
                first_detection = min(valid_starts)
                delay = first_detection - t_start
                delays.append(cls._convert_unit(delay, unit))

        if not delays:
            return np.nan

        values = [abs(i) for i in delays] if absolute else delays

        if reduction == "list":
            return values
        if reduction == "mean":
            return float(np.mean(values))
        if reduction == "median":
            return float(np.median(values))

        raise ValueError("reduction must be mean, median, or list")

    @staticmethod
    def _convert_unit(td, unit):
        seconds = td.total_seconds()

        if unit == "seconds":
            return seconds
        if unit == "minutes":
            return seconds / 60.0
        if unit == "hours":
            return seconds / 3600.0

        raise ValueError("unit must be seconds, minutes, or hours")

    @staticmethod
    def alerts_per_day(pred_events, timestamps, gap_threshold=None):
        timestamps = pd.DatetimeIndex(pd.to_datetime(timestamps)).sort_values()

        if len(timestamps) < 2:
            return 0.0

        diffs = timestamps.to_series().diff().dropna()

        if gap_threshold is None:
            gap_threshold = 10 * diffs.median()
        else:
            gap_threshold = pd.Timedelta(gap_threshold)

        covered = diffs[diffs <= gap_threshold].sum()
        duration_days = covered.total_seconds() / 86400

        if duration_days <= 0:
            return 0.0

        return len(pred_events) / duration_days

    @staticmethod
    def make_json_serializable(obj):
        if obj is None:
            return None
        if isinstance(obj, (str, int, float, bool)):
            return obj
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, (list, tuple)):
            return [MetricsBase.make_json_serializable(x) for x in obj]
        if isinstance(obj, dict):
            return {str(k): MetricsBase.make_json_serializable(v) for k, v in obj.items()}
        if isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        if isinstance(obj, pd.Timedelta):
            return str(obj)
        return str(obj)

    @staticmethod
    def format_model_params(obj):
        model_params = MetricsBase.make_json_serializable(obj)

        if isinstance(model_params, str):
            return model_params
        if model_params is None:
            return pd.NA

        return json.dumps(model_params, ensure_ascii=False)

    @staticmethod
    def change_pr(df, model_name, recall, precision, f1_anom):
        df.loc[model_name, "Recall"] = round(recall * 100, 2)
        df.loc[model_name, "Precision"] = round(precision * 100, 2)
        df.loc[model_name, "F1_anom"] = round(f1_anom * 100, 2)
        return df

    @staticmethod
    def plot_pr_curve(y_true, y_scores, model_name="model"):
        precision, recall, _ = precision_recall_curve(y_true, y_scores)
        pr_auc_value = average_precision_score(y_true, y_scores)

        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=recall,
                y=precision,
                mode="lines",
                name=f"{model_name} (PR-AUC={pr_auc_value:.4f})",
            )
        )
        fig.update_layout(
            title="Precision-Recall Curve",
            xaxis_title="Recall",
            yaxis_title="Precision",
            template="plotly_white",
        )
        fig.show()
        return pr_auc_value

    @staticmethod
    def rank_transform(scores, ascending=True):
        scores = pd.Series(scores)
        return scores.rank(method="average", pct=True, ascending=ascending)

    @staticmethod
    def calibrate_scores_by_val(val_scores_series, test_scores_series, higher_is_more_anomalous=True):
        val_scores_series = pd.Series(val_scores_series).dropna().sort_values().reset_index(drop=True)
        test_scores_series = pd.Series(test_scores_series)

        val_array = val_scores_series.to_numpy()
        test_array = test_scores_series.to_numpy()

        if len(val_array) == 0:
            return pd.Series(np.nan, index=test_scores_series.index)

        if higher_is_more_anomalous:
            calibrated = np.searchsorted(val_array, test_array, side="right") / len(val_array)
        else:
            calibrated = 1.0 - (np.searchsorted(val_array, test_array, side="left") / len(val_array))

        return pd.Series(calibrated, index=test_scores_series.index)


class ExperimentBase(MetricsBase, ABC):
    dataset_map = {
        1: "1",
        4: "4",
        6: "6",
        9: "9",
        10: "10",
        11: "11",
    }

    @staticmethod
    def patch_dagmm(dagmm_module, use_deterministic = False):
        def patched_ae_forward(self, x, return_latent=False):
            enc = self.encoder(x.reshape(x.shape[0], -1).float())
            dec = self.decoder(enc)
            recon_x = dec.reshape(x.shape)
            if return_latent:
                return recon_x, enc
            return recon_x

        def patched_dagmm_forward(self, x):
            dec, enc = self.autoencoder(x, return_latent=True)

            a = x.reshape(x.shape[0], -1)
            b = dec.reshape(dec.shape[0], -1)

            cos_distance = F.cosine_similarity(a, b, dim=1).unsqueeze(-1)
            euclidean_distance = ((a - b) ** 2).mean(dim=1).sqrt().unsqueeze(-1)

            z = torch.cat([enc, euclidean_distance, cos_distance], dim=1)
            gamma = self.estimation(z)

            return enc, dec, z, gamma

        dagmm_module.AEModule.forward = patched_ae_forward
        dagmm_module.DAGMMModule.forward = patched_dagmm_forward

    @staticmethod
    def set_seed(seed=42, use_deterministic=False):
        os.environ['PYTHONHASHSEED'] = str(seed)
        #os.environ["OMP_NUM_THREADS"] = "1"
        #os.environ["MKL_NUM_THREADS"] = "1"
        #os.environ["OPENBLAS_NUM_THREADS"] = "1"
        #os.environ["NUMEXPR_NUM_THREADS"] = "1"
        random.seed(seed)
        np.random.seed(seed)
        torch.manual_seed(seed)
        if torch.cuda.is_available():
            torch.cuda.manual_seed(seed)
            torch.cuda.manual_seed_all(seed)
        #torch.use_deterministic_algorithms(True)
        torch.backends.cudnn.deterministic = use_deterministic
        torch.backends.cudnn.benchmark = False

    @staticmethod
    def make_windows(df, y, window_size, step=1):
        values = df.to_numpy(dtype=np.float32)
        labels = y.reset_index(drop=True)
        X_windows, y_windows = [], []

        for start in range(0, len(df) - window_size + 1, step):
            end = start + window_size
            window = values[start:end]
            window = window.T
            X_windows.append(window)
            y_windows.append(labels.iloc[end - 1])

        return np.asarray(X_windows, dtype=np.float32), np.asarray(y_windows)
    
    @staticmethod
    def make_window_index(df, window_size, step=1):
        """
        Индекс для окон: берем timestamp последней точки окна.
        """
        return pd.to_datetime(df.index[window_size - 1::step])

    @staticmethod
    def get_feature_columns(df):
        """
        target — целевая переменная.
        is_failure — не обучаем, оставляем только для анализа/отрисовки.
        """
        return [c for c in df.columns if c not in ["target", "is_failure"]]

    @staticmethod
    def load_split_data(num_data, use_miss=True):
        train = pd.read_csv(f"../data_post/supervised/{num_data}/train.csv", index_col=0)
        val = pd.read_csv(f"../data_post/supervised/{num_data}/val.csv", index_col=0)

        test_paths = sorted(glob.glob(f"../data_post/supervised/{num_data}/test*.csv"))

        if len(test_paths) == 0 and os.path.exists(f"../data_post/supervised/{num_data}/test.csv"):
            test = pd.read_csv(f"../data_post/supervised/{num_data}/test.csv", index_col=0)

            return train, val, test

        tests = [pd.read_csv(path, index_col=0) for path in test_paths]

        return train, val, tests

    def resolve_event_params(self, train):
        lengths = self.get_anomaly_lengths_time(train)
        min_len = lengths["min"]
        median_len = lengths["median"]

        return self.event_params(min_len, median_len)

    @abstractmethod
    def prepare_scaled_data(self, num_data, scaler_cls, use_miss=True, use_scaling=True):
        pass

    @abstractmethod
    def fit_model(self, model, config, prepared, num_data, model_name_save):
        pass

    @abstractmethod
    def score_model(self, model, config, prepared, num_data, model_name_save):
        pass

    @abstractmethod
    def add_results(self, df, **kwargs):
        pass

    def compute_metrics(
        self,
        y_true,
        test_scores_series,
        val_scores_series,
        q,
        alpha,
        gap_fill,
        min_event,
        event_interval,
        use_calibrate=False,
        higher_is_more_anomalous=True,
        true_event_min_len=1,
        gap_threshold="3h",
    ):
        test_scores_series = pd.Series(test_scores_series).copy()
        val_scores_series = pd.Series(val_scores_series).copy()

        if use_calibrate:
            calibrated_val = self.calibrate_scores_by_val(
                val_scores_series=val_scores_series,
                test_scores_series=test_scores_series,
                higher_is_more_anomalous=higher_is_more_anomalous,
            )
            test_scores_series = self.calibrate_scores_by_val(
                val_scores_series=val_scores_series,
                test_scores_series=test_scores_series,
                higher_is_more_anomalous=higher_is_more_anomalous,
            )
            val_scores_series = calibrated_val

        test_scores_series.index = y_true.index

        if getattr(self, "score_mode", "proba") == "predict":
            threshold = np.nan
            y_pred_bin = test_scores_series.astype(int)
            y_pred = y_pred_bin.to_numpy()
        else:
            threshold = float(val_scores_series.quantile(q) * alpha)
            y_pred_bin = (test_scores_series > threshold).astype(int)
            y_pred = y_pred_bin.to_numpy()

        pred_events = self.binary_to_events(y_pred_bin, gap_minutes=gap_fill, min_len_minutes=min_event)
        true_events = self.binary_to_events(y_true, gap_minutes=gap_fill, min_len_minutes=true_event_min_len)

        event_r = self.event_recall(true_events, pred_events, max_min=event_interval)
        event_p = self.event_precision(true_events, pred_events, max_min=event_interval)
        f1_event = 0 if (event_p + event_r) == 0 else 2 * event_p * event_r / (event_p + event_r)

        alerts = self.alerts_per_day(pred_events, y_true.index, gap_threshold=gap_threshold)

        return {
            "threshold_value": threshold,
            "y_pred": y_pred,
            "y_pred_bin": y_pred_bin,
            "pred_events": pred_events,
            "true_events": true_events,
            "pr_auc_val": round(self.pr_auc(y_true, test_scores_series) * 100, 2),
            "f1_val": round(self.f1_at_fpr(y_true, test_scores_series, target_fpr=0.05) * 100, 2),
            "p100": round(self.precision_at_k(y_true, test_scores_series, k=100) * 100, 2),
            "r100": round(self.recall_at_k(y_true, test_scores_series, k=100) * 100, 2),
            "recall": round(recall_score(y_true, y_pred_bin, pos_label=1) * 100, 2),
            "precision": round(precision_score(y_true, y_pred_bin, pos_label=1) * 100, 2),
            "f1_anom": round(f1_score(y_true, y_pred, pos_label=1) * 100, 2),
            "event_r": round(event_r * 100, 2),
            "event_p": round(event_p * 100, 2),
            "f1_event": round(f1_event * 100, 2),
            "ttd_list": self.time_to_detect(true_events, pred_events, reduction="list", unit="minutes", max_min=event_interval),
            "ttd_mean": round(self.time_to_detect(true_events, pred_events, reduction="mean", unit="minutes", max_min=event_interval), 2),
            "ttd_median": round(self.time_to_detect(true_events, pred_events, reduction="median", unit="minutes", max_min=event_interval), 2),
            "alerts": round(alerts, 2),
            "cm": confusion_matrix(y_true, y_pred),
            "test_scores_series": test_scores_series,
            "val_scores_series": val_scores_series,
        }

    def save_results_sheet(
        self,
        num_data,
        model_name,
        metrics,
        model_params,
        train_time_sec,
        inference_time_sec,
        val_scores_series,
        q,
        alpha,
        file_path="../results_s.xlsx",
    ):
        sheet_name = str(num_data)

        if os.path.exists(file_path):
            try:
                results = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)
            except ValueError:
                results = pd.DataFrame()
        else:
            results = pd.DataFrame()

        if model_name in results.index:
            if train_time_sec is None and "train time sec" in results.columns:
                old_val = results.loc[model_name, "train time sec"]
                if pd.notna(old_val):
                    train_time_sec = float(old_val)

            if inference_time_sec is None and "inference time sec" in results.columns:
                old_val = results.loc[model_name, "inference time sec"]
                if pd.notna(old_val):
                    inference_time_sec = float(old_val)

            if model_params is None and "model params" in results.columns:
                old_params = results.loc[model_name, "model params"]
                if pd.notna(old_params):
                    model_params = old_params

        results = self.add_results(
            df=results,
            model_name=model_name,
            pr_auc=metrics["pr_auc_val"],
            f1_at_fpr_1=metrics["f1_val"],
            recall=metrics["recall"],
            precision=metrics["precision"],
            precision_at_100=metrics["p100"],
            recall_at_100=metrics["r100"],
            event_recall=metrics["event_r"],
            event_precision=metrics["event_p"],
            f1_event=metrics["f1_event"],
            ttd_mean=metrics["ttd_mean"],
            ttd_median=metrics["ttd_median"],
            alerts=metrics["alerts"],
            n_true_events=len(metrics["true_events"]),
            n_pred_events=len(metrics["pred_events"]),
            train_time_sec=train_time_sec,
            inference_time_sec=inference_time_sec,
            model_params=model_params,
            threshold=(
                "predict"
                if pd.isna(metrics["threshold_value"])
                else f'{q}, {"-" if val_scores_series.quantile(q) < 0 else "+"}, {alpha}'
            ),
            f1_anom=metrics["f1_anom"],
        )

        mode = "a" if os.path.exists(file_path) else "w"
        if mode == "a":
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                results.to_excel(writer, sheet_name=sheet_name, index=True)
        else:
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                results.to_excel(writer, sheet_name=sheet_name, index=True)

        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        ws.column_dimensions["A"].width = 35
        for i, col in enumerate(results.columns, start=2):
            max_len = results[col].astype(str).map(len).max() if len(results) > 0 else len(col)
            ws.column_dimensions[get_column_letter(i)].width = max(max_len, len(col)) + 2
        wb.save(file_path)
        wb.close()

    def save_summary_sheet(self, num_data, model_name, metrics, file_path="../results_all_s.xlsx"):
        dataset_name = self.dataset_map[num_data]

        sheet_to_metrics = {
            "PR_event": {
                "Event Recall": metrics["event_r"],
                "Event Precision": metrics["event_p"],
            },
            "PR-AUC, F1": {
                "F1 event": metrics["f1_event"],
                "PR-AUC": metrics["pr_auc_val"],
            },
        }

        for summary_sheet, metric_dict in sheet_to_metrics.items():
            if os.path.exists(file_path):
                try:
                    summary_df = pd.read_excel(file_path, sheet_name=summary_sheet, header=[0, 1], index_col=0)
                except ValueError:
                    summary_df = pd.DataFrame()
            else:
                summary_df = pd.DataFrame()

            for metric_name in metric_dict.keys():
                col = (dataset_name, metric_name)
                if col not in summary_df.columns:
                    summary_df[col] = pd.NA

            for metric_name, metric_value in metric_dict.items():
                summary_df.loc[model_name, (dataset_name, metric_name)] = metric_value

            mode = "a" if os.path.exists(file_path) else "w"
            if mode == "a":
                with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    summary_df.to_excel(writer, sheet_name=summary_sheet)
            else:
                with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                    summary_df.to_excel(writer, sheet_name=summary_sheet)

        wb = load_workbook(file_path)
        for summary_sheet in sheet_to_metrics.keys():
            ws = wb[summary_sheet]
            ws.column_dimensions["A"].width = 35
            for col_idx in range(2, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            ws.freeze_panes = "B3"
        wb.save(file_path)
        wb.close()

    @staticmethod
    def print_metrics(metrics):
        print("PR-AUC:", metrics["pr_auc_val"])
        print("Recall:", metrics["recall"])
        print("Precision:", metrics["precision"])
        print("F1_anom", metrics["f1_anom"])
        print("F1 @ FPR=5%:", metrics["f1_val"])
        print("Recall@100:", metrics["r100"])
        print("Precision@100:", metrics["p100"])
        print("Event Recall:", metrics["event_r"])
        print("Event Precision:", metrics["event_p"])
        print("F1 event:", metrics["f1_event"])
        print("Time-to-detect mean (minutes):", metrics["ttd_mean"])
        print("Time-to-detect median (minutes):", metrics["ttd_median"])
        print("Time-to-detect list (minutes):", metrics["ttd_list"])
        print("Alerts per day:", metrics["alerts"])
        print("n true events:", len(metrics["true_events"]))
        print("n pred events:", len(metrics["pred_events"]))
        print(metrics["cm"])

    @staticmethod
    def plot_events_and_scores(num_data, prepared, metrics):
        pred_events = metrics["pred_events"]
        true_events = metrics["true_events"]

        if num_data == 6 and len(pred_events) <= 100:
            fig = make_subplots(
                rows=2, cols=1, shared_xaxes=True,
                row_heights=[0.3, 0.7], vertical_spacing=0.05,
                specs=[[{"secondary_y": True}], [{}]]
            )

            for start, end in true_events:
                fig.add_trace(
                    go.Scatter(x=[start, end], y=[1, 1], mode="lines",
                               line=dict(width=12, color="green"),
                               showlegend=False, name="true_event"),
                    row=1, col=1, secondary_y=False
                )
                fig.add_vline(x=start, line_width=2, line_dash="dash",
                              line_color="green", opacity=0.7, row=1, col=1)

            for start, end in pred_events:
                fig.add_trace(
                    go.Scatter(x=[start, end], y=[0, 0], mode="lines",
                               line=dict(width=12, color="red"),
                               showlegend=False, name="pred_event"),
                    row=1, col=1, secondary_y=False
                )
                fig.add_vline(x=start, line_width=2, line_dash="dot",
                              line_color="red", opacity=0.7, row=1, col=1)

            fig.add_trace(
                go.Scatter(
                    x=metrics["test_scores_series"].index,
                    y=metrics["test_scores_series"],
                    line=dict(color="black", width=2),
                    name="score",
                ),
                row=1, col=1, secondary_y=True
            )

            df = prepared["test"].copy()
            fig.add_trace(go.Scatter(x=df.index, y=df.iloc[:, 0], line=dict(color="blue"), name=df.columns[0]), row=2, col=1)
            fig.add_trace(go.Scatter(x=df.index, y=df.iloc[:, 1], line=dict(color="orange"), name=df.columns[1]), row=2, col=1)

            fig.update_layout(template="plotly_white", height=600, title="Signals + Events + Score")
            fig.update_yaxes(tickvals=[0, 1], ticktext=["pred", "true"], row=1, col=1)
            fig.show()

        if len(pred_events) <= 100:
            fig = go.Figure()

            for start, end in true_events:
                fig.add_trace(go.Scatter(x=[start, end], y=[1, 1], mode="lines",
                                         line=dict(width=12, color="green"),
                                         name="true_event", showlegend=False))
                fig.add_vline(x=start, line_width=2, line_dash="dash", line_color="green", opacity=0.7)

            for start, end in pred_events:
                fig.add_trace(go.Scatter(x=[start, end], y=[0, 0], mode="lines",
                                         line=dict(width=12, color="red"),
                                         name="pred_event", showlegend=False))
                fig.add_vline(x=start, line_width=2, line_dash="dot", line_color="red", opacity=0.7)

            fig.update_layout(
                yaxis=dict(tickvals=[0, 1], ticktext=["pred", "true"]),
                title="True vs Predicted Events",
                template="plotly_white",
                height=300,
            )
            fig.show()

    def run(
        self,
        num_data,
        config,
        model_name_save,
        model_name,     
        q=0.95,
        alpha=1,
        model=None,
        save_result=True,
        train_model=False,
        use_miss=True,
        use_calibrate=False,
        higher_is_more_anomalous=True,
        use_scaling=True,
        scaler_cls=RobustScaler,
        use_deterministic=False
    ):
        prepared = self.prepare_scaled_data(
            num_data=num_data,
            scaler_cls=scaler_cls,
            use_miss=use_miss,
            use_scaling=use_scaling,
        )

        print(f'Размерности: train: {prepared["train"].shape}, val: {prepared["val"].shape}')
        print(
            f'Параметры исследования аномалий: максимальный пропуск: {prepared["gap_fill"]}, '
            f'минимальная длина аномалии: {prepared["min_event"]}, '
            f'интервал вокруг начала аномалии: {prepared["event_interval"]}, '
            f'размер окна: {prepared["window_size"]}'
        )
        print(f'Количество аномальных точек в датасете: {prepared["y_true"].sum()}')

        adapted_config = self.adapt_config(config, prepared)

        if train_model:
            model_train, train_time_sec = self.fit_model(
                model=model,
                config=adapted_config,
                prepared=prepared,
                num_data=num_data,
                model_name_save=model_name_save,
                use_deterministic=use_deterministic
            )
            _, val_scores_series, test_scores_series, inference_time_sec = self.score_model(
                model=model,
                config=adapted_config,
                prepared=prepared,
                num_data=num_data,
                model_name_save=model_name_save,
                model_instance=model_train,
            )
        else:
            model_train, val_scores_series, test_scores_series, inference_time_sec = self.score_model(
                model=model,
                config=adapted_config,
                prepared=prepared,
                num_data=num_data,
                model_name_save=model_name_save,
            )
            train_time_sec = None

        metrics = self.compute_metrics(
            y_true=prepared["y_true"],
            test_scores_series=test_scores_series,
            val_scores_series=val_scores_series,
            q=q,
            alpha=alpha,
            gap_fill=prepared["gap_fill"],
            min_event=prepared["min_event"],
            event_interval=prepared["event_interval"],
            use_calibrate=use_calibrate,
            higher_is_more_anomalous=higher_is_more_anomalous,
            true_event_min_len=self.true_event_min_len,
            gap_threshold=self.gap_threshold,
        )

        self.print_metrics(metrics)
        #self.plot_events_and_scores(num_data=num_data, prepared=prepared, metrics=metrics)

        if save_result:
            self.save_results_sheet(
                num_data=num_data,
                model_name=model_name,
                metrics=metrics,
                model_params=self.extract_model_params(model_train, adapted_config, train_model),
                train_time_sec=train_time_sec,
                inference_time_sec=inference_time_sec,
                val_scores_series=metrics["val_scores_series"],
                q=q,
                alpha=alpha,
            )
            self.save_summary_sheet(
                num_data=num_data,
                model_name=model_name,
                metrics=metrics,
            )

        return model_train, metrics["val_scores_series"], metrics["test_scores_series"], train_time_sec, inference_time_sec

    def adapt_config(self, config, prepared):
        return config

    def extract_model_params(self, model_train, adapted_config, train_model):
        return adapted_config if train_model else None


class MLPRunner(ExperimentBase):
    true_event_min_len = 1
    gap_threshold = "3h"

    def __init__(self, score_mode="proba"):
        if score_mode not in ["proba", "predict"]:
            raise ValueError("score_mode должен быть 'proba' или 'predict'")
        self.score_mode = score_mode

    def prepare_scaled_data(self, num_data, scaler_cls, use_miss, use_scaling):
        loaded = self.load_split_data(num_data=num_data, use_miss=use_miss)

        train = loaded[0].copy()
        val = loaded[1].copy()

        train.index = pd.to_datetime(train.index)
        val.index = pd.to_datetime(val.index)

        gap_fill, min_event, event_interval, window_size = self.resolve_event_params(train)

        feature_names = self.get_feature_columns(train)

        y_train_raw = train["target"].astype(int)
        y_val_raw = val["target"].astype(int)

        scaler = scaler_cls() if use_scaling else None

        if use_scaling:
            train_norm = pd.DataFrame(
                scaler.fit_transform(train[feature_names]),
                index=train.index,
                columns=feature_names,
            )
            val_norm = pd.DataFrame(
                scaler.transform(val[feature_names]),
                index=val.index,
                columns=feature_names,
            )
        else:
            train_norm = train[feature_names].copy()
            val_norm = val[feature_names].copy()

        X_train, y_train = self.make_windows(train_norm, y_train_raw, window_size)
        X_val, y_val = self.make_windows(val_norm, y_val_raw, window_size)

        result = {
            "train": train,
            "val": val,
            "X_train": X_train,
            "y_train": y_train,
            "X_val": X_val,
            "y_val": y_val,
            "val_index": self.make_window_index(val, window_size),
            "feature_names": feature_names,
            "gap_fill": gap_fill,
            "min_event": min_event,
            "event_interval": event_interval,
            "window_size": window_size,
            "scaler": scaler,
            "use_scaling": use_scaling,
        }

        if len(loaded) == 3 and isinstance(loaded[2], pd.DataFrame):
            test = loaded[2].copy()
            test.index = pd.to_datetime(test.index)

            y_test_raw = test["target"].astype(int)

            if use_scaling:
                test_norm = pd.DataFrame(
                    scaler.transform(test[feature_names]),
                    index=test.index,
                    columns=feature_names,
                )
            else:
                test_norm = test[feature_names].copy()

            X_test, y_test = self.make_windows(test_norm, y_test_raw, window_size)
            test_index = self.make_window_index(test, window_size)

            result.update({
                "test": test,
                "X_test": X_test,
                "y_test": y_test,
                "test_index": test_index,
                "y_true": pd.Series(y_test, index=test_index, name="target").astype(int),
            })

            if "is_failure" in test.columns:
                result["is_failure"] = test["is_failure"].reindex(test_index).astype(int)

        else:
            tests = [t.copy() for t in loaded[2]]

            X_test_list = []
            test_index_list = []
            y_true_parts = []
            is_failure_parts = []

            for test_part in tests:
                test_part.index = pd.to_datetime(test_part.index)

                y_test_raw = test_part["target"].astype(int)

                if use_scaling:
                    test_norm = pd.DataFrame(
                        scaler.transform(test_part[feature_names]),
                        index=test_part.index,
                        columns=feature_names,
                    )
                else:
                    test_norm = test_part[feature_names].copy()

                X_part, y_part = self.make_windows(test_norm, y_test_raw, window_size)
                idx_part = self.make_window_index(test_part, window_size)

                X_test_list.append(X_part)
                test_index_list.append(idx_part)

                y_true_parts.append(
                    pd.Series(y_part, index=idx_part, name="target").astype(int)
                )

                if "is_failure" in test_part.columns:
                    is_failure_parts.append(
                        test_part["is_failure"].reindex(idx_part).astype(int)
                    )

            result.update({
                "test": pd.concat(tests).sort_index(),
                "X_test_list": X_test_list,
                "test_index_list": test_index_list,
                "y_true": pd.concat(y_true_parts).sort_index(),
            })

            if is_failure_parts:
                result["is_failure"] = pd.concat(is_failure_parts).sort_index()

        print("Windowed shapes:")
        print("X_train:", result["X_train"].shape)
        print("y_train:", result["y_train"].shape)
        print("X_val:", result["X_val"].shape)
        print("y_val:", result["y_val"].shape)

        if "X_test" in result:
            print("X_test:", result["X_test"].shape)
            print("y_test:", result["y_test"].shape)
        else:
            for i, X_part in enumerate(result["X_test_list"], start=1):
                print(f"X_test{i}:", X_part.shape)

        return result

    def _get_scores(self, model_train, X):
        if self.score_mode == "predict":
            return np.asarray(model_train.predict(X)).astype(int)

        proba = np.asarray(model_train.predict_proba(X))

        if proba.ndim == 1:
            return proba

        classes = list(model_train.classes_)

        if 1 in classes:
            pos_idx = classes.index(1)
        elif "1" in classes:
            pos_idx = classes.index("1")
        else:
            pos_idx = 1

        return proba[:, pos_idx]

    @staticmethod
    def _suggest_from_param_grid(trial, param_grid):
        params = {}

        for name, values in param_grid.items():
            if isinstance(values, tuple) and len(values) == 2:
                low, high = values
                if isinstance(low, int) and isinstance(high, int):
                    params[name] = trial.suggest_int(name, low, high)
                else:
                    params[name] = trial.suggest_float(name, float(low), float(high))
            else:
                params[name] = trial.suggest_categorical(name, list(values))

        return params

    def _validation_objective_score(self, y_val, val_scores, val_index, q, alpha, gap_fill, min_event):
        y_val_series = pd.Series(y_val, index=val_index, name="target").astype(int)
        val_scores_series = pd.Series(val_scores, index=val_index, name="score")

        if len(y_val_series) == 0 or y_val_series.sum() == 0:
            return {
                "objective": 0.0,
                "f1_anom": 0.0,
                "pr_auc": 0.0,
                "alerts": 0.0,
                "threshold": np.nan,
            }

        pr_auc = float(average_precision_score(y_val_series, val_scores_series))
        threshold = float(val_scores_series.quantile(q) * alpha)
        y_pred = (val_scores_series > threshold).astype(int)
        f1_anom = float(f1_score(y_val_series, y_pred, pos_label=1, zero_division=0))

        pred_events = self.binary_to_events(
            y_pred,
            gap_minutes=gap_fill,
            min_len_minutes=min_event,
        )
        alerts = float(
            self.alerts_per_day(
                pred_events,
                val_index,
                gap_threshold=self.gap_threshold,
            )
        )

        objective = 0.0 if alerts <= 0 else float(f1_anom * pr_auc / alerts)

        return {
            "objective": objective,
            "f1_anom": f1_anom,
            "pr_auc": pr_auc,
            "alerts": alerts,
            "threshold": threshold,
        }

    def make_optuna_objective(self, prepared, param_grid, base_config=None, model=None, q=0.95, alpha=1.0):
        X_train = prepared["X_train"]
        y_train = prepared["y_train"]
        X_val = prepared["X_val"]
        y_val = prepared["y_val"]
        val_index = prepared["val_index"]
        gap_fill = prepared["gap_fill"]
        min_event = prepared["min_event"]

        model_cls = MLPClassifier if model is None else model
        param_grid = dict(param_grid or {})
        base_config = dict(base_config or {})

        def objective(trial):
            trial_params = self._suggest_from_param_grid(trial, param_grid)

            trial_q = trial_params.pop("q", q)
            trial_alpha = trial_params.pop("alpha", alpha)

            config = dict(base_config)
            config.update(trial_params)

            model_train = model_cls(**config)
            model_train.fit(X_train, y_train)

            val_scores = self._get_scores(model_train, X_val)
            score_info = self._validation_objective_score(
                y_val=y_val,
                val_scores=val_scores,
                val_index=val_index,
                q=trial_q,
                alpha=trial_alpha,
                gap_fill=gap_fill,
                min_event=min_event,
            )

            trial.set_user_attr("config", config)
            trial.set_user_attr("q", trial_q)
            trial.set_user_attr("alpha", trial_alpha)
            trial.set_user_attr("f1_anom", score_info["f1_anom"])
            trial.set_user_attr("pr_auc", score_info["pr_auc"])
            trial.set_user_attr("alerts", score_info["alerts"])
            trial.set_user_attr("threshold", score_info["threshold"])

            return score_info["objective"]

        return objective

    def tune_hyperparameters(
        self,
        num_data,
        param_grid,
        base_config=None,
        model=None,
        n_trials=50,
        q=0.95,
        alpha=1.0,
        use_miss=True,
        use_scaling=True,
        scaler_cls=RobustScaler,
        study_name=None,
        direction="maximize",
        show_progress_bar=True,
    ):

        prepared = self.prepare_scaled_data(
            num_data=num_data,
            scaler_cls=scaler_cls,
            use_miss=use_miss,
            use_scaling=use_scaling,
        )

        objective = self.make_optuna_objective(
            prepared=prepared,
            param_grid=param_grid,
            base_config=base_config,
            model=model,
            q=q,
            alpha=alpha,
        )

        study = optuna.create_study(
            direction=direction,
            study_name=study_name,
        )
        study.optimize(
            objective,
            n_trials=n_trials,
            show_progress_bar=show_progress_bar,
        )

        return study

    def fit_model(self, model, config, prepared, num_data, model_name_save, use_deterministic=False):
        self.set_seed(42, use_deterministic=use_deterministic)

        model_train = MLPClassifier(**config) if model is None else model(**config)

        start_train = time.perf_counter()

        print("Train")
        model_train.fit(prepared["X_train"], prepared["y_train"])

        train_time_sec = time.perf_counter() - start_train

        os.makedirs("../models/supervised", exist_ok=True)

        model_train.save(f"../models/supervised/{model_name_save}{num_data}")

        return model_train, train_time_sec

    def score_model(self, model, config, prepared, num_data, model_name_save, model_instance=None):
        if model_instance is None:
            model_train = model.load_from_path(Path(f"../models/supervised/{model_name_save}{num_data}.zip"))
        else:
            model_train = model_instance

        start_infer = time.perf_counter()

        print("Predict val")
        val_scores = self._get_scores(model_train, prepared["X_val"])

        val_scores_series = pd.Series(
            val_scores,
            index=prepared["val_index"],
            name="score",
        ).sort_index()

        print("Predict test")

        if "X_test_list" in prepared:
            test_scores_parts = []

            for X_part, idx_part in zip(prepared["X_test_list"], prepared["test_index_list"]):
                scores_part = self._get_scores(model_train, X_part)

                test_scores_parts.append(
                    pd.Series(scores_part, index=idx_part, name="score")
                )

            test_scores_series = pd.concat(test_scores_parts).sort_index()

        else:
            test_scores = self._get_scores(model_train, prepared["X_test"])

            test_scores_series = pd.Series(
                test_scores,
                index=prepared["test_index"],
                name="score",
            ).sort_index()

        inference_time_sec = time.perf_counter() - start_infer

        return model_train, val_scores_series, test_scores_series, inference_time_sec

    def add_results(self, df, **kwargs):
        row = {
            "PR-AUC": kwargs["pr_auc"],
            "F1@FPR5%": kwargs["f1_at_fpr_1"],
            "Recall": kwargs["recall"],
            "Precision": kwargs["precision"],
            "F1_anom": kwargs["f1_anom"],
            "Precision@100": kwargs["precision_at_100"],
            "Recall@100": kwargs["recall_at_100"],
            "Event Recall": kwargs["event_recall"],
            "Event Precision": kwargs["event_precision"],
            "F1 event": kwargs["f1_event"],
            "TTD mean": kwargs["ttd_mean"],
            "TTD median": kwargs["ttd_median"],
            "Alerts per day": kwargs["alerts"],
            "n true events": kwargs["n_true_events"],
            "n pred events": kwargs["n_pred_events"],
            "threshold": kwargs["threshold"],
            "train time sec": round(kwargs["train_time_sec"], 2)
            if kwargs["train_time_sec"] is not None
            else pd.NA,
            "inference time sec": round(kwargs["inference_time_sec"], 2)
            if kwargs["inference_time_sec"] is not None
            else pd.NA,
            "model params": self.format_model_params(kwargs["model_params"]),
        }

        if df.empty:
            df = pd.DataFrame(columns=row.keys())

        for col in row:
            if col not in df.columns:
                df[col] = pd.NA

        df.loc[kwargs["model_name"]] = row

        return df